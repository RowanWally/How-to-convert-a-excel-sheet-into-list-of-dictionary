import re
import os
import shutil
import logging
from xlsxwriter import Workbook
from sys import stdout
from multiprocessing.pool import ThreadPool 
import datetime
import xml.etree.ElementTree as ET
# Non-Standard Libraries
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from texttable import Texttable 
date_strftime_format = "%d-%b-%y %H:%M:%S"
message_format = "%(asctime)s || %(levelname)s || %(message)s"
logging.basicConfig(format= message_format, datefmt= date_strftime_format, stream=stdout , level=logging.INFO )

def convert_excel_to_dict_backend_file():
    wb = load_workbook(filename = f'file/file.xlsx' , data_only=True)
    sheet = wb['host'] # this sheet contains the physical devices management IPs
    rd=list(sheet.iter_rows(values_only=True))
    rd_enhanced=[]
    for i in rd:
        row=[]
        row.append(i[1])
        row.append(i[3])
        rd_enhanced.append(row)
        
    rd_lis_dict= []
        
    for j in rd_enhanced:
        data={}
        data["Device Name"]=j[0]
        data["IP Address"]=j[1]
        rd_lis_dict.append(data)
    #print(rd_lis_dict[1:4])
    return rd_lis_dict
    
    def convert_excel_to_dict_another():
    wb = load_workbook(filename = f'file2/file2.xlsx' , data_only=True)
    sheet = wb['Ark'] # this sheet contains the physical devices management IPs
    rd=list(sheet.iter_rows(values_only=True))
    rd_enhanced=[]
    for i in rd:
        row=[]
        row.append(i[3])
        row.append(i[4])
        rd_enhanced.append(row)    
    rd_lis_dict= []
    for j in cyberark_enhanced:
        data={}
        data["IP Address"]=j[0]
        data["Username"]=j[1]
        rd_lis_dict.append(data)
    return rd_lis_dict
    
    def main():
    out1=[]
    ALL=[]
    ALLL=[]
    exetra=[]
    out2=[]
    output1=[]
    out1=convert_excel_to_dict_backend_file()
    out2=convert_excel_to_dict_another()
    for j in out1:
       for i in anotherone:
           if i["Device Name"] == j["Device Name"]:
                data={}
                data["Device Name"]=j["Device Name"]
                data["IP Address"]=j["IP Address"]
                data["AZ"]=i["AZ"]
                data["Role"]=i["Role"]
                output1.append(data)
               
                
    for j in out1: 
       l=0
       for k in output1:
                if j["Device Name"] == k["Device Name"]:
                    l=1
                    break
       if l==1:
          continue
       data={}
       data["Device Name"]=j["Device Name"]
       data["IP Address"]=j["IP Address"]
       data["AZ"]=" "
       data["Role"]="Empty_string"
       exetra.append(data)

   ALL= outpu1+ exetra

    for y in out2:
        for x in ALL:
            if x["IP Address"] == y["IP Address"]:
                try:
                   if not re.search(r"svc|root",y["Username"]):
                        final={}
                        final["Device Name"]=x["Device Name"]
                        final["IP Address"]=x["IP Address"]
                        final["Role"]=x["Role"]
                        final["Username"]=y["Username"]
                        ALLL.append(final)
                except:
                    pass
                    
    ordered_list=["Device Name", "IP Address", "Role", "Username"]  
    wb=Workbook("final.xlsx")    
    ws=wb.add_worksheet("all")
    first_row=0
    for header in ordered_list:
        col=ordered_list.index(header) # We are keeping order.
        ws.write(first_row,col,header) # We have written first row which is the header of worksheet also.

    row=1
    
    for player in ALLL:
        for _key,_value in player.items():
            col=ordered_list.index(_key)
            ws.write(row,col,_value)
        row+=1 #enter the next row
        
    wb.close()         
                    
    
     

if __name__ == "__main__":
    main()
   
    
